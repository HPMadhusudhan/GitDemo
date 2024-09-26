from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

chromeoptions=webdriver.ChromeOptions()
chromeoptions.add_experimental_option("detach",True)
driver=webdriver.Chrome(options=chromeoptions)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.implicitly_wait(5)
fruit_name="Apple"
updatedprice="1000"

# click on download button
# driver.find_element(By.ID,"downloadButton").click()

# Edit excel sheet
book=openpyxl.load_workbook("C:\\Users\\Admin\\Downloads\\download.xlsx")
sheet=book.active


for j in range(1, sheet.max_column+1):
    if sheet.cell(row=1,column=j).value=="fruit_name":
        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=j).value==fruit_name:
                fruitrow=i
                print(fruitrow)
    if sheet.cell(row=1, column=j).value=="price":
        pricecolumn=j
        print(pricecolumn)

sheet.cell(row=fruitrow,column=pricecolumn).value=updatedprice

# After update, you need to save file to see changes are reflected
book.save("C:\\Users\\Admin\\Downloads\\download.xlsx")



# Upload excel sheet
fileadd=driver.find_element(By.ID,"fileinput")
#to upload a file make sure that element has type="file" attribute
# or else selenium will not be able to upload file without that attribute
fileadd.send_keys("C:\\Users\\Admin\\Downloads\\download.xlsx")

#explicit wait
WebDriverWait(driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH,"//div[@class='Toastify__toast-body']/div[2]")))
print(driver.find_element(By.XPATH,"//div[@class='Toastify__toast-body']/div[2]").text)

#checking updated value on the table in webpag
# transversing back to parent tag--> /parent::parent_tagname
# smart way of parameterizing by get the column number of "Price" column
price_column=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")

# Pass fruit name and column number dynamically
actual_price=driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
print(actual_price)

assert actual_price==updatedprice